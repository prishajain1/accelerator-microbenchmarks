import argparse
import os
import re
import jsonlines
import pandas as pd
from google.cloud import storage
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict
import tempfile

METRICS_TO_REPORT = [
    "ici_bandwidth_gbyte_s_p50",
    "ici_bandwidth_gbyte_s_p90",
    "ici_bandwidth_gbyte_s_p95",
    "ici_bandwidth_gbyte_s_p99",
    "ici_bandwidth_gbyte_s_avg",
]

def get_num_chips(tpu_type: str) -> int:
    """Extracts the number of chips from the TPU type string."""
    match = re.search(r"-(\d+)$", tpu_type)
    if match:
        return int(match.group(1))
    raise ValueError(f"Could not extract number of chips from TPU type: {tpu_type}")

def download_blob(bucket_name, source_blob_name, destination_file_name):
    """Downloads a blob from the bucket."""
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(source_blob_name)
    blob.download_to_filename(destination_file_name)
    # print(f"Blob {source_blob_name} downloaded to {destination_file_name}.")

def upload_blob(bucket_name, source_file_name, destination_blob_name):
    """Uploads a file to the bucket."""
    storage_client = storage.Client()
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_filename(source_file_name)
    print(f"File {source_file_name} uploaded to {destination_blob_name}.")

def list_blobs(bucket_name, prefix):
    """Lists all the blobs in the bucket that start with the prefix."""
    storage_client = storage.Client()
    blobs = storage_client.list_blobs(bucket_name, prefix=prefix)
    return [blob.name for blob in blobs]

def parse_gcs_path(gcs_path):
    """Parses GCS path into bucket name and prefix."""
    if not gcs_path.startswith("gs://"):
        raise ValueError("GCS path must start with gs://")
    parts = gcs_path[5:].split("/", 1)
    bucket_name = parts[0]
    prefix = parts[1] if len(parts) > 1 else ""
    return bucket_name, prefix

def get_dimension_config(benchmark_name):
    """
    Returns the key to be used as the primary dimension for row labels
    and a function to extract a sortable value from it.
    """
    # TODO: Customize this function based on the benchmarks
    DIM_KEYS = ['m', 'n', 'dim', 'size', 'buffer_size', 'dimension']
    # Default dimension key
    return DIM_KEYS, lambda x: x

def identify_dimension_column(df, potential_dim_keys):
    """Identifies the column in the DataFrame that varies the most, likely the dimension."""
    if df.empty:
        return None

    for key in potential_dim_keys:
        if key in df.columns:
             # Check if this column has more than one unique value
            if df[key].nunique() > 1:
                is_numeric = pd.api.types.is_numeric_dtype(df[key])
                return key, is_numeric
            # If only one unique value, it might still be the dimension if only one row exists
            if df.shape[0] == 1:
                is_numeric = pd.api.types.is_numeric_dtype(df[key])
                return key, is_numeric

    # Fallback or more complex inference
    for col in df.columns:
        if col not in METRICS_TO_REPORT and df[col].nunique() > 1:
             is_numeric = pd.api.types.is_numeric_dtype(df[col])
             return col, is_numeric
    return None, False

def format_sheet(sheet, num_chips):
    """Applies formatting to the sheet."""
    # Bold headers
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for i, metric in enumerate(METRICS_TO_REPORT):
        col_start = 1 + i * 3
        # Metric name header (A1, D1, G1, ...)
        cell = sheet.cell(row=1, column=col_start)
        cell.font = bold_font
        cell.alignment = center_alignment
        sheet.merge_cells(start_row=1, start_column=col_start, end_row=1, end_column=col_start + 1)

        # Sub headers (A2, B2, D2, E2, ...)
        cell = sheet.cell(row=2, column=col_start) # dimensions\TPUs
        cell.font = bold_font
        cell = sheet.cell(row=2, column=col_start + 1) # Num chips
        cell.font = bold_font
        cell.alignment = center_alignment

    # Adjust column widths
    for column_cells in sheet.columns:
        try:
            length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2
        except ValueError:
            pass


def generate_excel_report(gcs_run_dir: str, tpu_type: str):
    """Generates the Excel report from JSONL files in GCS."""
    bucket_name, run_prefix = parse_gcs_path(gcs_run_dir)
    num_chips = get_num_chips(tpu_type)
    xlml_prefix = os.path.join(run_prefix, "xlml/")

    print(f"Scanning for .jsonl files in {bucket_name}/{xlml_prefix}")
    blob_names = list_blobs(bucket_name, xlml_prefix)
    jsonl_files = [b for b in blob_names if b.endswith(".jsonl")]

    if not jsonl_files:
        print("No .jsonl files found.")
        return

    benchmark_data = defaultdict(list)
    temp_dir = tempfile.mkdtemp()

    for file_blob in jsonl_files:
        try:
            # microbenchmark_all_gather_2024-08-26T18:20:59.215844+00:00Z.jsonl
            file_name = os.path.basename(file_blob)
            parts = file_name.split("_")
            if len(parts) < 3: continue
            benchmark_name = "_".join(parts[1:-1])

            local_file = os.path.join(temp_dir, file_name)
            download_blob(bucket_name, file_blob, local_file)

            with jsonlines.open(local_file) as reader:
                for obj in reader:
                    flat_data = obj.get("metadata", {})
                    flat_data.update(obj.get("metrics", {}))
                    benchmark_data[benchmark_name].append(flat_data)
        except Exception as e:
            print(f"Error processing file {file_blob}: {e}")

    if not benchmark_data:
        print("No data parsed from jsonl files.")
        return

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    for benchmark_name, data in benchmark_data.items():
        print(f"Processing benchmark: {benchmark_name}")
        sheet = wb.create_sheet(title=benchmark_name)
        df = pd.DataFrame(data)

        if df.empty:
            print(f"No data for {benchmark_name}")
            continue

        potential_dim_keys, dim_extractor = get_dimension_config(benchmark_name)
        dim_col, is_numeric = identify_dimension_column(df, potential_dim_keys)

        if not dim_col:
            print(f"Could not identify dimension column for {benchmark_name}")
            continue

        print(f"Using '{dim_col}' as dimension column for {benchmark_name}")

        # Sort by dimension column
        if is_numeric:
            df[dim_col] = pd.to_numeric(df[dim_col])
        df = df.sort_values(by=dim_col)

        dimensions = df[dim_col].tolist()

        for i, metric in enumerate(METRICS_TO_REPORT):
            if metric not in df.columns:
                print(f"Metric {metric} not found in data for {benchmark_name}")
                continue

            col_start = 1 + i * 3
            # Write headers
            sheet.cell(row=1, column=col_start, value=metric)
            sheet.cell(row=2, column=col_start, value=f"{dim_col}\\TPUs")
            sheet.cell(row=2, column=col_start + 1, value=num_chips)

            # Write data
            for row_idx, dim_val in enumerate(dimensions):
                sheet.cell(row=row_idx + 3, column=col_start, value=dim_val)
                metric_val = df.iloc[row_idx][metric]
                sheet.cell(row=row_idx + 3, column=col_start + 1, value=metric_val)

        format_sheet(sheet, num_chips)

    # Save workbook to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
        wb.save(tmp_excel.name)
        excel_file_path = tmp_excel.name

    # Upload to GCS
    excel_blob_name = os.path.join(run_prefix, "benchmark_report.xlsx")
    upload_blob(bucket_name, excel_file_path, excel_blob_name)

    # Clean up temp files
    os.remove(excel_file_path)
    for root, dirs, files in os.walk(temp_dir, topdown=False):
        for name in files:
            os.remove(os.path.join(root, name))
        for name in dirs:
            os.rmdir(os.path.join(root, name))
    os.rmdir(temp_dir)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel report from benchmark results in GCS.")
    parser.add_argument(
        "--gcs_run_dir",
        type=str,
        required=True,
        help="GCS directory containing the run results (e.g., gs://bucket/path/to/run).",
    )
    parser.add_argument(
        "--tpu_type",
        type=str,
        required=True,
        help="TPU type used for the benchmark (e.g., v5p-256).",
    )
    args = parser.parse_args()
    generate_excel_report(args.gcs_run_dir, args.tpu_type)
